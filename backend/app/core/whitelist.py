"""
MindGuard Institutional Whitelist Roster
Pre-authorized emails and their designated roles for self-service registration.
Supports both direct dictionary roster and dynamic synchronization with Excel (.xlsx) roster files.
"""

import os
import glob
from typing import Optional, Dict
from app.models.users import UserRole

# -------------------------------------------------------------------------
# Base / Built-in Institutional Whitelist Roster
# -------------------------------------------------------------------------
BASE_AUTHORIZED_ROSTER: Dict[str, UserRole] = {
    # 1. NMIMS Student Roster (from Excel and institutional list)
    "makkena.lahari06@nmims.in": UserRole.STUDENT,
    "siripuramvaishnavi.goud20@nmims.in": UserRole.STUDENT,
    "aashritha.reddy35@nmims.in": UserRole.STUDENT,
    "ratnachand.kancharla04@nmims.in": UserRole.STUDENT,
    "viraj.meedintisunny69@nmims.in": UserRole.STUDENT,
    "pabbashashank.goud048@nmims.edu.in": UserRole.STUDENT,
    "vsreshta.reddy42@nmims.in": UserRole.STUDENT,
    "sherymounika.reddy32@nmims.in": UserRole.STUDENT,
    "gumudavelli.vikram22@nmims.in": UserRole.STUDENT,
    "baleeshwar.badam31@nmims.in": UserRole.STUDENT,
    "ptarunkumar.reddy43@nmims.in": UserRole.STUDENT,
    "kusunurilakshmi.ramya36@nmims.in": UserRole.STUDENT,
    "naguboyina.divya59@nmims.in": UserRole.STUDENT,
    "kuchurusai.krishna34@nmims.in": UserRole.STUDENT,
    "pravalika.kolluri70@nmims.in": UserRole.STUDENT,
    "canil.kumar65@nmims.in": UserRole.STUDENT,
    "lam.jahnavy23@nmims.in": UserRole.STUDENT,
    "kottakapu.janshi64@nmims.in": UserRole.STUDENT,
    "md.rayyan33@nmims.in": UserRole.STUDENT,
    "ashi.sharma85@nmims.in": UserRole.STUDENT,
    "pasupulasai.teja37@nmims.in": UserRole.STUDENT,
    "chinthakunta.harini40@nmims.in": UserRole.STUDENT,
    "avuti.anoushka30@nmims.in": UserRole.STUDENT,
    "brungishiva.ganesh21@nmims.in": UserRole.STUDENT,
    "t.rishikesh.talpalikar41@nmims.in": UserRole.STUDENT,
    "s.saikarthik.reddy18@nmims.in": UserRole.STUDENT,
    "sathwika.sv54@nmims.in": UserRole.STUDENT,
    "student@nmims.in": UserRole.STUDENT,
    "student1@nmims.in": UserRole.STUDENT,
    "student2@nmims.in": UserRole.STUDENT,
    "student3@nmims.in": UserRole.STUDENT,

    # 2. NMIMS Clinical Counselor Staff Roster
    "naresh.vurukonda@nmims.edu": UserRole.COUNSELOR,
    "chandrakant.wani@nmims.edu": UserRole.COUNSELOR,
    "vinayak.mukkawar@nmims.edu": UserRole.COUNSELOR,
    "rahul.koshti@nmims.edu": UserRole.COUNSELOR,
    "nikita.pande@nmims.edu": UserRole.COUNSELOR,
    "bhanusree.y@nmims.edu": UserRole.COUNSELOR,
    "counselor@nmims.edu": UserRole.COUNSELOR,
    "counselor1@nmims.edu": UserRole.COUNSELOR,
    "counselor2@nmims.edu": UserRole.COUNSELOR,
    "dr.singh@nmims.edu": UserRole.COUNSELOR,
    "dr.kapoor@nmims.edu": UserRole.COUNSELOR,
    "wellness.counselor@nmims.edu": UserRole.COUNSELOR,

    # 3. NMIMS Institutional Administration Roster
    "raja.govindaacharyk@nmims.edu": UserRole.ADMIN,
    "admin@nmims.edu": UserRole.ADMIN,
    "superadmin@nmims.edu": UserRole.ADMIN,
    "dean.studentaffairs@nmims.edu": UserRole.ADMIN,
}

_EXCEL_ROSTER_CACHE: Dict[str, UserRole] = {}
_LAST_MTIME: float = 0.0

def _find_excel_roster_path() -> Optional[str]:
    """Find the institutional Excel roster file in common locations."""
    search_candidates = [
        r"d:\mindguard\nmims emails.xlsx",
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "nmims emails.xlsx")),
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "nmims emails.xlsx")),
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "nmims emails.xlsx")),
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "nmims emails.xlsx")),
    ]
    for path in search_candidates:
        if os.path.isfile(path):
            return path
    return None

def _load_excel_roster() -> Dict[str, UserRole]:
    """Dynamically reads and parses emails and roles from the Excel spreadsheet."""
    global _EXCEL_ROSTER_CACHE, _LAST_MTIME
    excel_path = _find_excel_roster_path()
    if not excel_path:
        return {}

    try:
        current_mtime = os.path.getmtime(excel_path)
        if _EXCEL_ROSTER_CACHE and current_mtime == _LAST_MTIME:
            return _EXCEL_ROSTER_CACHE

        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        roster: Dict[str, UserRole] = {}

        # Scan headers
        header_map: Dict[int, UserRole] = {}
        for col_idx in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col_idx).value
            if not val:
                continue
            val_str = str(val).lower().strip()
            if "student" in val_str:
                header_map[col_idx] = UserRole.STUDENT
            elif "counselor" in val_str:
                header_map[col_idx] = UserRole.COUNSELOR
            elif "admin" in val_str:
                header_map[col_idx] = UserRole.ADMIN

        # Scan rows
        for row_idx in range(2, sheet.max_row + 1):
            for col_idx, assigned_role in header_map.items():
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    email_clean = str(cell_val).strip().lower()
                    if "@" in email_clean:
                        roster[email_clean] = assigned_role

        _EXCEL_ROSTER_CACHE = roster
        _LAST_MTIME = current_mtime
        return roster
    except Exception:
        return _EXCEL_ROSTER_CACHE or {}

def get_authorized_roster() -> Dict[str, UserRole]:
    """Returns the combined roster from base list and live Excel file."""
    combined = dict(BASE_AUTHORIZED_ROSTER)
    excel_data = _load_excel_roster()
    combined.update(excel_data)
    return combined

def get_authorized_role(email: str) -> Optional[UserRole]:
    """
    Returns the designated UserRole if the email is on the institutional roster, else None.
    """
    normalized_email = email.lower().strip()
    roster = get_authorized_roster()
    return roster.get(normalized_email)

def is_email_whitelisted(email: str) -> bool:
    """
    Returns True if the email is present in the institutional whitelist roster.
    """
    return get_authorized_role(email) is not None
